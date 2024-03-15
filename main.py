from __future__ import annotations

import re
from datetime import datetime, timedelta
from functools import wraps

import telebot  # pyTelegramBotAPI
import xlrd  # read old version .xlsx files
from openpyxl import Workbook  # create and save .xlsx file

from settings import token, users_id

#  time settings
today = datetime.now()
tomorrow = datetime.now() + timedelta(days=1)
DATE_TO = tomorrow.strftime("%d.%m")
DATE_CALL = today.strftime("%d.%m")


# EXCEL ################################################################################################################

def _get_car_number(string: str) -> str:
    """ Find and return car number in str. """
    pattern = r"\w{1,2}\d{3}\w{0,2}\d{2,3}"  # example: А123ВС799 or АВ12377
    match = re.search(pattern, string)
    return match.group() if match else ""


def get_vehicle_inspection_time(column: int, sheet: Workbook.active) -> str:
    """ Return srt time from column. """
    value = sheet.cell_value
    hour = int(value(1, column)) if value(1, column) != 42 else int(value(1, column - 1))
    minute = int(value(2, column))
    return f'{hour:02d}:{minute:02d}'


def get_time_car_list(sheet: Workbook.active) -> list[list[str]]:
    """ """
    result_list = []

    for column in range(2, sheet.ncols - 1):
        # get time
        time = get_vehicle_inspection_time(column, sheet)

        # get car
        rows_with_car_numbers = (3, 4, 5, 6)
        for row in rows_with_car_numbers:
            string = sheet.cell_value(row, column)  # sheet --> sh
            if string != 42:  # xlrd print "42" in empty cell
                result_car = _get_car_number(string)
                result_list.append([time, result_car])

    return result_list


def write_vehicle_inspection_driver_table(time_car_list: list[list[str]], name: str) -> None:
    # create file
    out_book = Workbook()
    sheet = out_book.active

    # create table
    for rw in range(1, len(time_car_list) + 1):
        sheet.cell(column=1, row=rw).value = DATE_TO
        sheet.cell(column=2, row=rw).value = time_car_list[rw - 1][0]
        sheet.cell(column=3, row=rw).value = time_car_list[rw - 1][1]
        sheet.cell(column=4, row=rw).value = ''
        sheet.cell(column=5, row=rw).value = ''
        sheet.cell(column=6, row=rw).value = name
        sheet.cell(column=7, row=rw).value = DATE_CALL

    # save table
    out_book.save("result.xlsx")
    out_book.close()


# TELEBOT ##############################################################################################################

bot = telebot.TeleBot(token)
xlsx_file = None
csv_files = []


def private_access():
    """ decorate for user verification. """

    def deco_restrict(func):
        @wraps(func)
        def f_restrict(message, *args, **kwargs):
            user = message.from_user.id
            if user in users_id:
                return func(message, *args, **kwargs)
            else:
                bot.reply_to(message, text=f'Who are you? Keep on walking...')

        return f_restrict  # true decorator

    return deco_restrict


@bot.message_handler(commands=['help'])
@private_access()
def send_welcome(message):
    bot.reply_to(message, """\
Привет! Бот конвертирует список авто из 1с в список для обзвона.
Просто пришлите xlsx файл из 1с мне.\
""")


# Проверка даты ТО
def check_date(file, message):
    book = xlrd.open_workbook(file)
    sheet = book.sheet_by_index(0)
    # check_correct_date
    data_in_file = str((sheet.cell_value(0, 2)).split()[0])
    if data_in_file != tomorrow.strftime("%d.%m.%Y"):
        bot.send_message(message.chat.id,
                         "‼️❗️‼️❗️‼️❗️‼️\nДата ТО в файле и завтрашняя дата не совпадают. "
                         "Проверьте правильно ли скачан файл")


# Функция для сохранения файла на сервере
def save_file(message, filename):
    file_info = bot.get_file(message.document.file_id)
    downloaded_file = bot.download_file(file_info.file_path)
    with open(filename, 'wb') as new_file:
        new_file.write(downloaded_file)
    return filename


def clear_files():
    global xlsx_file, csv_files
    xlsx_file = None
    csv_files.clear()


# Handle sent .xlsx and .csv documents
@bot.message_handler(func=lambda message: True, content_types=['document'])
@private_access()
def default_command(message):
    global xlsx_file, csv_files

    # Прием XLSX, проверка что файл только один
    if message.document.mime_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
        if xlsx_file is None:
            xlsx_file = save_file(message, 'file.xlsx')
            bot.send_message(message.chat.id, 'XLSX done.')
        else:
            bot.send_message(message.chat.id, '❗️Отправлено несколько XLSX - файлов.')
            clear_files()
            return

    # Прием csv
    elif message.document.mime_type == 'text/csv':
        filename = f'file_{len(csv_files) + 1}.csv'  # Используем уникальное имя файла
        csv_files.append(save_file(message, filename))
        bot.send_message(message.chat.id, 'CSV done.')


# Функция для обработки файлов
def process_files():
    global xlsx_file, csv_files
    if xlsx_file is not None and len(csv_files) > 0:
        # Прочитать XLSX-файл и обработать его
        print('process_files success', xlsx_file)

        # Обработать CSV-файлы
        for csv_file in csv_files:
            print(23, csv_file)
            # Здесь можно выполнить дополнительные операции с данными

        # Очистить переменные после обработки файлов
        xlsx_file = None
        csv_files.clear()
    else:
        print('Недостаточно файлов для обработки. \n Попробуйте загрузить еще раз')




    clear_files()


# Обработчик команды /process
@bot.message_handler(commands=['process', 'start'])
@private_access()
def process_message(message):
    process_files()
    # сделать


    bot.send_message(message.chat.id, 'Файлы обработаны.')


@bot.message_handler(commands=['clear'])
@private_access()
def process_message(message):
    clear_files()
    bot.send_message(message.chat.id, 'Clear done!')


def run_bot():
    bot.polling()


def stop_bot():
    clear_files()
    bot.stop_polling()


if __name__ == "__main__":
    run_bot()
